"""
@author: Sai kumar Dandla
"""
import numpy as np
import cv2
from keras_facenet import FaceNet
embedder = FaceNet()
from pil import Image
from matplotlib import pyplot
from numpy import asarray
import datetime
import os



import pandas as pd
import openpyxl
from openpyxl import load_workbook

trail = embedder.extract('images//' + 'sai (1).jpg', threshold=0.95)

def show_face(filename,box,name,info,required_size=(160, 160)):
    # load image from file
    image = Image.open(filename)
	# convert to RGB, if needed
    image = image.convert('RGB')
	# convert to array
    pixels = asarray(image)
    
    x1,y1,width,height = box

    x2, y2 = x1 + width, y1 + height
        
    # extract the face
    face = pixels[y1:y2, x1:x2]
    # resize pixels to the model size
    image = Image.fromarray(face)
#    image = image.resize(required_size)
    face_array = asarray(image)
    
    pyplot.figure("Please Close this Window")
    pyplot.title(name + '\n Date & Time ' + info)
    pyplot.imshow(face_array)    
    pyplot.show()   
    
    return


def runit():
    try:
        detections = embedder.extract('image_0.png', threshold=0.95)
    
        face_info = detections[0]
        face_embedding = face_info['embedding']
        box = face_info['box']
        face_embedding=face_embedding.reshape(1,-1)
        
        import facetrain
        name_dict = np.load('name_dict.npy',allow_pickle='TRUE').item()
    #    names = np.load('names.npy')
    #    names = names.tolist()
        model = facetrain.model
        a = int(model.predict(face_embedding))
        print(name_dict[a])
        
        
    
        now= datetime.datetime.now()
        date = datetime.datetime.now().date()
        time = datetime.datetime.now().time()
        print (now.strftime("%d-%m-%Y  %H:%M:%S"))
        show_face('image_0.png',box,name_dict[a],now.strftime("%d-%m-%Y  %H:%M:%S"))
        
        
        df = pd.DataFrame({'Name': [name_dict[a]],'Date': [date.strftime("%d-%m-%Y")],'Time': [time.strftime("%H:%M:%S")]})
        folder = 'Attendence//'
        sheet_name = folder + "Attendence_{}.xlsx".format(date.strftime("%d-%m-%Y"))
        writer = pd.ExcelWriter(sheet_name, engine='openpyxl')
        writer.book = load_workbook(sheet_name)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        reader = pd.read_excel(sheet_name)
        df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
        writer.close()
        print ("\nAttendence Stored\n\n")
    except IndexError:
        pass
    return


now= datetime.datetime.now()
date = datetime.datetime.now().date()
folder = 'Attendence//'
sheet_name = folder + "Attendence_{}.xlsx".format(date.strftime("%d-%m-%Y"))

if not os.path.isfile(sheet_name):
    
    wb = openpyxl.Workbook()    
    wb.save(sheet_name)    
    df = pd.DataFrame({'Name': [' '],'Date': [' '],'Time': [' ']})
    df.to_excel(sheet_name,index=False)  


cam = cv2.VideoCapture(0)
cv2.namedWindow("Press Spacebar for Attendence")
img_counter = 0



while True:
    ret, frame = cam.read()
    if not ret:
        print("failed to grab frame")
        break
    cv2.imshow("Press Spacebar for Attendence", frame)

    k = cv2.waitKey(1)
    if k%256 == 27:
        # ESC pressed
        print("Escape , closing...\n\n")
        break
    elif k%256 == 32:
        # SPACE pressed
        img_name = "image_{}.png".format(img_counter)
        cv2.imwrite(img_name, frame)
        print("{} written!\n".format(img_name))
#        img_counter += 1
        runit()
        
#        import time
#        time.sleep(3)
        
#        l = cv2.waitKey(3000)
##        if l%256 == 32:
#        pyplot.close()
#    elif k%256 == 8:
#        pyplot.close()
      

cam.release()
cv2.destroyAllWindows()




